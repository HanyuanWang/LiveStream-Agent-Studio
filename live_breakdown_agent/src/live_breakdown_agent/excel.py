from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import EventRow, validate_event_rows


def export_excel(rows: list[EventRow], output_path: Path) -> Path:
    validate_event_rows(rows)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "直播逐字稿拆解"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.append(["时间戳", "事件", "逐字稿"])

    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    sheet.row_dimensions[1].height = 26

    for row in rows:
        sheet.append([row.timestamp, row.event, row.transcript])
        current = sheet.max_row
        for cell in sheet[current]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        sheet.cell(current, 1).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        if row.is_key_event:
            sheet.cell(current, 2).font = Font(color="C00000", bold=True)
        sheet.row_dimensions[current].height = 46

    for index, width in enumerate((22, 34, 100), 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    workbook.save(output_path)
    return output_path
