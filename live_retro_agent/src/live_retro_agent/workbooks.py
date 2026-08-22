from __future__ import annotations

import csv
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .config import Settings


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).lstrip("\t").strip()


def extract_xlsx(path: Path, target: Path, settings: Settings) -> dict[str, Any]:
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        raise RuntimeError(f"无法读取 Excel：{path.name}\n{exc}") from exc
    try:
        sheets = [
            {"name": sheet.title, "rows": [list(row) for row in sheet.iter_rows(values_only=True)]}
            for sheet in workbook.worksheets
        ]
    finally:
        workbook.close()
    return {"source": str(path), "sheets": sheets}


def read_csv(path: Path) -> list[list[Any]]:
    for encoding in ("utf-8-sig", "gb18030", "utf-8"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                return [row for row in csv.reader(handle)]
        except UnicodeDecodeError:
            continue
    raise RuntimeError(f"无法识别 CSV 编码：{path.name}")


def normalize_label(value: Any) -> str:
    return re.sub(r"[\s\n\r()（）/\-_]+", "", clean(value)).lower()


def find_header(rows: list[list[Any]], required_groups: list[list[str]], limit: int = 60) -> tuple[int, dict[str, int]] | None:
    aliases = [[normalize_label(x) for x in group] for group in required_groups]
    for idx, row in enumerate(rows[:limit]):
        normalized = [normalize_label(x) for x in row]
        mapping: dict[str, int] = {}
        ok = True
        for group, candidates in zip(required_groups, aliases):
            found = next((col for col, value in enumerate(normalized) if value and any(c in value or value in c for c in candidates)), None)
            if found is None:
                ok = False
                break
            mapping[group[0]] = found
        if ok:
            return idx, mapping
    return None


def iter_sheets(data: dict[str, Any]) -> Iterable[tuple[str, list[list[Any]]]]:
    for sheet in data.get("sheets", []):
        yield clean(sheet.get("name")), sheet.get("rows") or []


def parse_number(value: Any) -> float:
    text = clean(value).replace(",", "").replace("¥", "").replace("￥", "")
    if not text or text in {"-", "--", "None", "缺失"}:
        return 0.0
    multiplier = 1.0
    if text.endswith("万"):
        multiplier, text = 10000.0, text[:-1]
    elif text.endswith("w") or text.endswith("W"):
        multiplier, text = 10000.0, text[:-1]
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group()) * multiplier if match else 0.0


def parse_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    text = clean(value)
    if not text:
        return None
    if text.endswith("Z") and "T" in text:
        try:
            return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
        except ValueError:
            pass
    patterns = (
        "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M", "%Y-%m-%dT%H:%M:%S", "%H:%M:%S", "%H:%M",
    )
    for pattern in patterns:
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue
    match = re.search(r"(20\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2}).*?(\d{1,2})[:时-](\d{1,2})(?:[:分](\d{1,2}))?", text)
    if match:
        y, mo, d, h, mi, s = [int(x or 0) for x in match.groups()]
        return datetime(y, mo, d, h, mi, s)
    return None


def parse_hms(value: Any) -> float | None:
    text = clean(value)
    matches = re.findall(r"(?<!\d)(\d{1,2}):(\d{2})(?::(\d{2}(?:\.\d+)?))?", text)
    if not matches:
        return None
    h, m, s = matches[0]
    if s == "":
        return int(h) * 60 + int(m)
    return int(h) * 3600 + int(m) * 60 + float(s)


def hms(seconds: float | None) -> str:
    if seconds is None:
        return "缺失"
    seconds = max(0, int(round(seconds)))
    hour, remain = divmod(seconds, 3600)
    minute, second = divmod(remain, 60)
    return f"{hour:02}:{minute:02}:{second:02}"
