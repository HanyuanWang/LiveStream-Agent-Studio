import csv
import io
import unittest

from openpyxl import Workbook

from live_scout_agent.importers import parse_leaderboard, parse_number


class ImporterTests(unittest.TestCase):
    def test_parse_chinese_units_and_ranges(self):
        self.assertEqual(parse_number("8.3万"), 83_000)
        self.assertEqual(parse_number("1.2亿"), 120_000_000)
        self.assertEqual(parse_number("10万-20万"), 150_000)

    def test_parse_csv_aliases(self):
        stream = io.StringIO()
        writer = csv.writer(stream)
        writer.writerow(["达人昵称", "粉丝数", "直播销售额", "GPM", "主页链接"])
        writer.writerow(["主播甲", "8.3万", "20万-30万", "1250", "https://www.douyin.com/user/abc"])
        rows, warnings = parse_leaderboard("榜单.csv", stream.getvalue().encode("utf-8-sig"))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["anchor_name"], "主播甲")
        self.assertEqual(rows[0]["followers"], 83_000)
        self.assertIsNone(rows[0]["estimated_gmv"])
        self.assertEqual(rows[0]["estimated_gmv_text"], "20万-30万")
        self.assertEqual(rows[0]["gpm"], 1250)

    def test_parse_xlsx(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["主播", "类目", "预估销量", "直播时长"])
        sheet.append(["主播乙", "女装", 1200, "02:30:00"])
        payload = io.BytesIO()
        workbook.save(payload)
        rows, _ = parse_leaderboard("蝉妈妈榜单.xlsx", payload.getvalue())
        self.assertEqual(rows[0]["anchor_name"], "主播乙")
        self.assertEqual(rows[0]["sales_volume"], 1200)
        self.assertAlmostEqual(rows[0]["duration_hours"], 2.5)

    def test_parse_chanmama_creator_goods_headers(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["蝉妈妈带货达人榜"])
        sheet.append(["排行", "达人", "抖音号", "直播销售额(元)", "直播销量(件)", "粉丝数", "带货类目", "直播场次"])
        sheet.append([1, "主播丙", "douyin-003", "50w~100w", "1.5w", "8.6w", "服饰内衣", 3])
        payload = io.BytesIO()
        workbook.save(payload)
        rows, warnings = parse_leaderboard("蝉妈妈带货达人榜.xlsx", payload.getvalue())
        self.assertEqual(warnings, ["未找到部分评分字段：gpm、uv_value、profile_url"])
        self.assertEqual(rows[0]["anchor_name"], "主播丙")
        self.assertIsNone(rows[0]["estimated_gmv"])
        self.assertEqual(rows[0]["estimated_gmv_text"], "50w~100w")
        self.assertEqual(rows[0]["sales_volume"], 15000)
        self.assertEqual(rows[0]["sales_volume_text"], "1.5w")
        self.assertEqual(rows[0]["followers"], 86000)
        self.assertEqual(rows[0]["sessions_7d"], 3)

    def test_keeps_sales_ranges_separate_from_sales_indexes(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "达人",
                "直播销售额(元)",
                "销售额指数",
                "直播销量(件)",
                "销量指数",
            ]
        )
        sheet.append(["主播指数测试", "1000w+", 5_609_284, "2.5w~5w", 7_309])
        payload = io.BytesIO()
        workbook.save(payload)
        rows, _ = parse_leaderboard("蝉妈妈_网页榜单.xlsx", payload.getvalue())
        row = rows[0]
        self.assertIsNone(row["estimated_gmv"])
        self.assertEqual(row["estimated_gmv_text"], "1000w+")
        self.assertEqual(row["gmv_index"], 5_609_284)
        self.assertIsNone(row["sales_volume"])
        self.assertEqual(row["sales_volume_text"], "2.5w~5w")
        self.assertEqual(row["sales_index"], 7_309)

    def test_keeps_chanmama_creator_detail_hyperlink(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["达人", "粉丝数"])
        sheet.append(["主播丁", "3.2w"])
        sheet["A2"].hyperlink = "https://www.chanmama.com/bloggerRank/example.html"
        payload = io.BytesIO()
        workbook.save(payload)
        rows, _ = parse_leaderboard("蝉妈妈达人榜.xlsx", payload.getvalue())
        self.assertEqual(
            rows[0]["analysis_url"],
            "https://www.chanmama.com/bloggerRank/example.html",
        )


if __name__ == "__main__":
    unittest.main()
