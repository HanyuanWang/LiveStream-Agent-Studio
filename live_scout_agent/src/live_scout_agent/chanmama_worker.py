from __future__ import annotations

import argparse
import json
import re
import time
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from playwright.sync_api import Download, Page, sync_playwright

from .chanmama import CHANMAMA_CDP_PORT, read_state, write_state


def safe_file_name(value: str) -> str:
    value = re.sub(r'[<>:"/\\|?*]+', "_", value).strip(" .")
    return value or f"蝉妈妈带货达人榜_{int(time.time())}.xlsx"


def page_details(page: Page) -> dict[str, Any]:
    try:
        return {"page_url": page.url, "page_title": page.title()}
    except Exception:
        return {}


def capture_page_map(page: Page, state_path: Path, name: str = "page_map.json") -> Path:
    map_path = state_path.with_name(name)
    elements = page.locator("a,button,input,[role='menuitem'],[role='tab']").evaluate_all(
        """elements => elements.slice(0, 300).map((element, index) => ({
            index,
            tag: element.tagName.toLowerCase(),
            text: (element.innerText || element.getAttribute('aria-label') || '').trim().slice(0, 160),
            href: element.getAttribute('href') || '',
            placeholder: element.getAttribute('placeholder') || '',
            role: element.getAttribute('role') || ''
        })).filter(item => item.text || item.href || item.placeholder)"""
    )
    body_text = page.locator("body").inner_text(timeout=10_000)[:30_000]
    payload = {**page_details(page), "body_text": body_text, "elements": elements}
    import json
    map_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return map_path


def capture_creator_links(page: Page, state_path: Path) -> Path:
    link_path = state_path.with_name("leaderboard_links.json")
    rows = page.locator("a[href*='/bloggerRank/'][href$='.html']").evaluate_all(
        """elements => elements.map(element => ({
            name: (element.innerText || '').trim(),
            href: element.href || element.getAttribute('href') || ''
        })).filter(item => item.name && item.href)"""
    )
    import json
    mapping = {item["name"]: item["href"] for item in rows}
    link_path.write_text(json.dumps(mapping, ensure_ascii=False, indent=2), encoding="utf-8")
    return link_path


def capture_visible_leaderboard(page: Page) -> dict[str, Any]:
    result = page.locator(".el-table, table").evaluate_all(
        """containers => {
            const normalize = value => {
                const text = typeof value === 'string'
                    ? value
                    : ((value && (value.innerText || value.textContent)) || '');
                return text.replace(/\\s+/g, ' ').trim();
            };
            const lineValues = element => (element.innerText || '')
                .split(/\\r?\\n/)
                .map(normalize)
                .filter(Boolean);
            for (const container of containers) {
                const isElementTable = container.classList &&
                    container.classList.contains('el-table');
                let headerElements = Array.from(
                    container.querySelectorAll(
                        isElementTable ? '.el-table__header thead th' : 'thead th'
                    )
                );
                if (!headerElements.length) {
                    headerElements = Array.from(container.querySelectorAll('th'));
                }
                const headers = headerElements.map(normalize);
                if (!headers.some(value => value.includes('达人')) ||
                    !headers.some(value => value.includes('直播销售额'))) {
                    continue;
                }
                const rowSelector = isElementTable
                    ? '.el-table__body-wrapper tbody tr'
                    : 'tbody tr';
                const rows = Array.from(container.querySelectorAll(rowSelector)).map(row => {
                    const cells = Array.from(row.querySelectorAll('td'));
                    const values = cells.map(cell => lineValues(cell));
                    const creatorIndex = headers.findIndex(value => value.includes('达人'));
                    const creatorCell = cells[creatorIndex];
                    const creatorLines = values[creatorIndex] || [];
                    const detailLink = creatorCell
                        ? Array.from(creatorCell.querySelectorAll('a'))
                            .map(anchor => anchor.href || anchor.getAttribute('href') || '')
                            .find(href => href.includes('chanmama.com') || href.includes('/bloggerRank/')) || ''
                        : '';
                    return {
                        values,
                        creator_name: creatorLines[0] || '',
                        douyin_id: creatorLines[1] || '',
                        detail_link: detailLink
                    };
                }).filter(item => item.creator_name);
                if (rows.length) {
                    return { headers, rows };
                }
            }
            return { headers: [], rows: [] };
        }"""
    )
    headers = result.get("headers") or []
    source_rows = result.get("rows") or []
    if not headers or not source_rows:
        structure = page.locator(
            "table, .el-table, [class*='table'], [class*='rank-list']"
        ).evaluate_all(
            """elements => elements.slice(0, 40).map(element => ({
                tag: element.tagName,
                class_name: typeof element.className === 'string' ? element.className : '',
                th_count: element.querySelectorAll('th').length,
                tr_count: element.querySelectorAll('tr').length,
                text: (element.innerText || '').replace(/\\s+/g, ' ').trim().slice(0, 300)
            }))"""
        )
        raise RuntimeError(
            "当前榜单表格没有可读取的数据；页面结构="
            + json.dumps(structure, ensure_ascii=False)[:4000]
        )

    def index_of(text: str) -> int:
        return next((index for index, header in enumerate(headers) if text in header), -1)

    indices = {
        "rank": index_of("排行"),
        "gmv": index_of("直播销售额"),
        "sales": index_of("直播销量"),
        "price": index_of("销售客单价"),
        "followers": index_of("粉丝数"),
        "category": index_of("带货类目"),
        "sessions": index_of("直播场次"),
    }

    def cell_value(item: dict[str, Any], key: str, *, prefer_last: bool = False) -> str:
        index = indices[key]
        values = item.get("values") or []
        lines = values[index] if 0 <= index < len(values) else []
        if not lines:
            return ""
        return str(lines[-1] if prefer_last else lines[0])

    rows = []
    for position, item in enumerate(source_rows, start=1):
        rank = cell_value(item, "rank") or str(position)
        rows.append(
            [
                rank,
                str(item.get("creator_name") or ""),
                str(item.get("douyin_id") or ""),
                cell_value(item, "gmv"),
                cell_value(item, "gmv", prefer_last=True),
                cell_value(item, "sales"),
                cell_value(item, "sales", prefer_last=True),
                cell_value(item, "price"),
                cell_value(item, "followers"),
                cell_value(item, "category"),
                cell_value(item, "sessions"),
                str(item.get("detail_link") or ""),
            ]
        )
    return {
        "columns": [
            "排行",
            "达人",
            "抖音号",
            "直播销售额(元)",
            "销售额指数",
            "直播销量(件)",
            "销量指数",
            "销售客单价",
            "粉丝数",
            "带货类目",
            "直播场次",
            "蝉妈妈详情页",
        ],
        "rows": rows,
    }


def build_local_leaderboard_xlsx(
    payload: dict[str, Any],
    output_path: Path,
    state_path: Path,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "达人榜单"
    columns = list(payload.get("columns") or [])
    rows = list(payload.get("rows") or [])
    sheet.append([payload.get("title") or "蝉妈妈带货达人榜"])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(columns), 1))
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="0B6B4F")
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.append([payload.get("subtitle") or ""])
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(columns), 1))
    sheet.append(columns)
    for cell in sheet[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17825F")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        if isinstance(row, dict):
            sheet.append([row.get(column, "") for column in columns])
        else:
            sheet.append(list(row))
    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(columns, start=1):
        values = [str(column)] + [str(sheet.cell(row=row, column=index).value or "") for row in range(4, sheet.max_row + 1)]
        sheet.column_dimensions[sheet.cell(row=3, column=index).column_letter].width = min(max(max(map(len, values)) + 2, 10), 36)
    workbook.save(output_path)
    input_path = state_path.with_name("visible_leaderboard.json")
    preview_path = state_path.with_name("visible_leaderboard_preview.png")
    input_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return preview_path


def first_visible(locator):
    for index in range(locator.count()):
        candidate = locator.nth(index)
        try:
            if candidate.is_visible():
                return candidate
        except Exception:
            continue
    return None


def save_export_diagnostics(
    page: Page,
    state_path: Path,
    events: dict[str, list[str]],
    button=None,
) -> Path:
    path = state_path.with_name("export_diagnostics.json")
    button_info: dict[str, Any] = {}
    event_listener_info: dict[str, Any] = {}
    if button is not None:
        try:
            button_info = button.evaluate(
                """element => {
                    const style = getComputedStyle(element);
                    const ancestors = [];
                    let current = element;
                    for (let level = 0; current && level < 8; level += 1) {
                        const currentStyle = getComputedStyle(current);
                        ancestors.push({
                            level,
                            tag: current.tagName,
                            id: current.id || '',
                            class_name: typeof current.className === 'string' ? current.className : '',
                            role: current.getAttribute('role') || '',
                            cursor: currentStyle.cursor,
                            pointer_events: currentStyle.pointerEvents,
                            onclick: current.getAttribute('onclick') || '',
                            outer_html: current.outerHTML.slice(0, 2500)
                        });
                        current = current.parentElement;
                    }
                    return {
                        tag: element.tagName,
                        text: (element.innerText || element.textContent || '').trim(),
                        outer_html: element.outerHTML.slice(0, 3000),
                        disabled: Boolean(element.disabled),
                        aria_disabled: element.getAttribute('aria-disabled') || '',
                        pointer_events: style.pointerEvents,
                        display: style.display,
                        visibility: style.visibility,
                        opacity: style.opacity
                        ,
                        parent_outer_html: element.parentElement
                            ? element.parentElement.outerHTML.slice(0, 5000)
                            : '',
                        ancestors
                    };
                }"""
            )
        except Exception as exc:
            button_info = {"inspection_error": str(exc)}
        try:
            cdp = page.context.new_cdp_session(page)
            remote = cdp.send(
                "Runtime.evaluate",
                {
                    "expression": (
                        "Array.from(document.querySelectorAll('button'))"
                        ".find(element => (element.innerText || '').trim() === '导出数据')"
                    ),
                    "returnByValue": False,
                },
            )
            object_id = remote.get("result", {}).get("objectId")
            if object_id:
                listeners = cdp.send(
                    "DOMDebugger.getEventListeners",
                    {"objectId": object_id, "depth": 4, "pierce": True},
                ).get("listeners", [])
                event_listener_info = {
                    "count": len(listeners),
                    "listeners": [
                        {
                            "type": item.get("type"),
                            "use_capture": item.get("useCapture"),
                            "passive": item.get("passive"),
                            "once": item.get("once"),
                            "script_id": item.get("scriptId"),
                            "line_number": item.get("lineNumber"),
                            "column_number": item.get("columnNumber"),
                            "handler": (item.get("handler") or {}).get("description", "")[:1000],
                        }
                        for item in listeners
                    ],
                }
            cdp.detach()
        except Exception as exc:
            event_listener_info = {"inspection_error": str(exc)}
    payload = {
        **page_details(page),
        "button": button_info,
        "event_listeners": event_listener_info,
        "events": events,
        "body_text": page.locator("body").inner_text(timeout=10_000)[:20_000],
    }
    import json
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def run(args: argparse.Namespace) -> int:
    profile_dir = Path(args.profile_dir)
    download_dir = Path(args.download_dir)
    state_path = Path(args.state_path)
    stop_path = Path(args.stop_path)
    profile_dir.mkdir(parents=True, exist_ok=True)
    download_dir.mkdir(parents=True, exist_ok=True)
    stop_path.unlink(missing_ok=True)
    downloaded: list[Path] = []
    initial_logged_in = bool(read_state(state_path).get("logged_in"))

    def update(**values: Any) -> None:
        state = read_state(state_path)
        state.update(values)
        write_state(state_path, state)

    def capture_download(download: Download) -> None:
        suggested = safe_file_name(download.suggested_filename)
        if Path(suggested).suffix.lower() not in {".xlsx", ".xlsm", ".csv", ".txt"}:
            suggested += ".xlsx"
        destination = download_dir / suggested
        if destination.exists():
            destination = download_dir / f"{destination.stem}_{int(time.time())}{destination.suffix}"
        download.save_as(destination)
        downloaded.append(destination)
        update(
            phase="downloaded",
            message=f"榜单已下载：{destination.name}",
            download_path=str(destination),
            busy=False,
            imported=False,
        )

    try:
        with sync_playwright() as playwright:
            attached_browser = None
            context = None
            # When the user leaves the dedicated login Chrome open, the
            # persistent profile is legitimately locked. Reuse that browser
            # through its loopback-only CDP endpoint instead of starting a
            # conflicting second Chrome process.
            try:
                attached_browser = playwright.chromium.connect_over_cdp(
                    f"http://127.0.0.1:{CHANMAMA_CDP_PORT}", timeout=3_000
                )
                if attached_browser.contexts:
                    context = attached_browser.contexts[0]
            except Exception:
                attached_browser = None
            if context is None:
                try:
                    context = playwright.chromium.launch_persistent_context(
                        str(profile_dir),
                        channel="chrome",
                        headless=False,
                        accept_downloads=True,
                        downloads_path=str(download_dir),
                        no_viewport=True,
                        args=["--disable-background-timer-throttling"],
                    )
                except Exception as exc:
                    message = str(exc)
                    if "Target page, context or browser has been closed" in message:
                        raise RuntimeError(
                            "蝉妈妈专用Chrome仍被旧会话占用。请关闭单独打开的蝉妈妈Chrome窗口，"
                            "再点击“打开专用浏览器登录”重新登录一次；之后即使窗口保持打开，"
                            "Agent也可以直接复用该窗口更新榜单。"
                        ) from exc
                    raise
            pages = context.pages
            page = pages[0] if pages else context.new_page()
            diagnostic_events: dict[str, list[str]] = {
                "console_errors": [],
                "page_errors": [],
                "request_failures": [],
                "dialogs": [],
                "requests_after_click": [],
                "responses_after_click": [],
                "check_right_response": [],
            }
            page.on(
                "console",
                lambda message: diagnostic_events["console_errors"].append(message.text)
                if message.type == "error"
                else None,
            )
            page.on("pageerror", lambda error: diagnostic_events["page_errors"].append(str(error)))
            page.on(
                "requestfailed",
                lambda request: diagnostic_events["request_failures"].append(
                    f"{request.method} {request.url}: {request.failure}"
                ),
            )
            def handle_dialog(dialog) -> None:
                diagnostic_events["dialogs"].append(f"{dialog.type}: {dialog.message}")
                dialog.accept()
            page.on("dialog", handle_dialog)
            page.on("download", capture_download)
            if not page.url or page.url == "about:blank":
                page.goto(args.start_url, wait_until="domcontentloaded", timeout=60_000)

            if args.mode == "login":
                update(
                    phase="waiting_for_login",
                    message="请在独立窗口中登录蝉妈妈，完成后回到控制台点击“我已完成登录”",
                    mode="login",
                    logged_in=False,
                    busy=True,
                    **page_details(page),
                )
            elif args.mode == "export":
                page.goto(
                    "https://www.chanmama.com/bloggerRank/liveBlogger/",
                    wait_until="domcontentloaded",
                    timeout=60_000,
                )
                # Chanmama may render the target route briefly and only then
                # redirect an expired session to register/login. Give that
                # client-side redirect enough time to settle before treating
                # the page as an authenticated leaderboard.
                page.wait_for_timeout(8000)
                if any(
                    marker in page.url.lower()
                    for marker in ("/register", "/login", "passport.chanmama")
                ):
                    update(
                        phase="not_configured",
                        message=(
                            "蝉妈妈登录状态已失效。请点击“打开专用浏览器登录”，"
                            "重新登录后再更新榜单。"
                        ),
                        logged_in=False,
                        busy=False,
                        mode="",
                        **page_details(page),
                    )
                    if attached_browser is None:
                        context.close()
                    return 2
                category = args.category if args.category else "全部"
                period_label = {"day": "日榜", "week": "周榜", "month": "月榜"}[args.period]
                if category != "全部":
                    page.get_by_text(category, exact=True).first.click(timeout=15_000)
                    page.wait_for_timeout(1800)
                page.get_by_text(period_label, exact=True).first.click(timeout=15_000)
                page.wait_for_timeout(1800)
                update(
                    phase="exporting",
                    message=f"已进入带货达人榜，正在导出“{category} · {period_label}”",
                    mode="export",
                    busy=True,
                    theme_id=args.theme_id,
                    category=category,
                    period=args.period,
                    **page_details(page),
                )
                creator_links_path = capture_creator_links(page, state_path)
                update(creator_links_path=str(creator_links_path))
                local_capture_succeeded = False
                try:
                    local_data = capture_visible_leaderboard(page)
                    local_data.update(
                        {
                            "title": f"蝉妈妈带货达人榜｜{category}｜{period_label}",
                            "subtitle": (
                                "由主播发现 Agent 直接读取当前网页榜单生成；"
                                "未使用蝉妈妈官方导出功能"
                            ),
                        }
                    )
                    destination = download_dir / safe_file_name(
                        f"蝉妈妈_{category}_{period_label}_网页榜单.xlsx"
                    )
                    preview_path = build_local_leaderboard_xlsx(
                        local_data, destination, state_path
                    )
                    downloaded.append(destination)
                    local_capture_succeeded = True
                    update(
                        phase="downloaded",
                        message=(
                            f"已直接读取网页榜单中的{len(local_data['rows'])}位达人，"
                            "无需使用蝉妈妈付费导出功能"
                        ),
                        download_path=str(destination),
                        preview_path=str(preview_path),
                        capture_method="visible_web_table",
                        busy=False,
                        imported=False,
                    )
                except Exception as exc:
                    if any(
                        marker in page.url.lower()
                        for marker in ("/register", "/login", "passport.chanmama")
                    ):
                        update(
                            phase="not_configured",
                            message=(
                                "蝉妈妈登录状态已失效。请点击“打开专用浏览器登录”，"
                                "重新登录后再更新榜单。"
                            ),
                            logged_in=False,
                            busy=False,
                            mode="",
                            **page_details(page),
                        )
                        if attached_browser is None:
                            context.close()
                        return 2
                    update(local_capture_error=str(exc))
                export_button = first_visible(page.locator("button:has-text('导出数据')"))
                if local_capture_succeeded:
                    pass
                elif export_button is None:
                    export_button = first_visible(page.locator("a:has-text('导出数据')"))
                if export_button is None:
                    export_button = first_visible(page.locator("[role='button']:has-text('导出数据')"))
                if export_button is None:
                    export_button = first_visible(page.locator(".el-button:has-text('导出数据')"))
                if export_button is None:
                    export_button = first_visible(page.get_by_role("button", name="导出数据", exact=True))
                if export_button is None:
                    export_button = first_visible(page.get_by_text("导出数据", exact=True))
                if export_button is not None:
                    try:
                        tag_name = export_button.evaluate("element => element.tagName.toLowerCase()")
                        if tag_name not in {"button", "a"}:
                            inner_button = first_visible(export_button.locator("button, a, [role='button']"))
                            if inner_button is not None:
                                export_button = inner_button
                            else:
                                export_button = export_button.locator(
                                    "xpath=ancestor::*["
                                    "self::button or self::a or @role='button' or "
                                    "contains(concat(' ', normalize-space(@class), ' '), ' el-button ')"
                                    "][1]"
                                )
                    except Exception:
                        pass
                if export_button is None:
                    debug_map = capture_page_map(page, state_path, "page_map_export_blocked.json")
                    screenshot_path = state_path.with_name("export_blocked.png")
                    page.screenshot(path=str(screenshot_path), full_page=True)
                    update(
                        phase="needs_action",
                        message="已进入目标榜单，但没有找到可点击的“导出数据”。窗口将保持打开，请检查会员提示或页面弹窗。",
                        busy=True,
                        debug_map=str(debug_map),
                        screenshot_path=str(screenshot_path),
                        **page_details(page),
                    )
                else:
                    try:
                        right_check: dict[str, Any] = {}
                        export_button.scroll_into_view_if_needed(timeout=10_000)
                        diagnostics_path = save_export_diagnostics(
                            page, state_path, diagnostic_events, export_button
                        )
                        diagnostic_events["requests_after_click"].clear()
                        diagnostic_events["responses_after_click"].clear()
                        page.on(
                            "request",
                            lambda request: diagnostic_events["requests_after_click"].append(
                                f"{request.method} {request.url}"
                            )
                            if "chanmama.com" in request.url
                            else None,
                        )
                        page.on(
                            "response",
                            lambda response: diagnostic_events["responses_after_click"].append(
                                f"{response.status} {response.request.method} {response.url}"
                            )
                            if "chanmama.com" in response.url
                            else None,
                        )
                        try:
                            with page.expect_response(
                                lambda response: "/export/common/checkRight" in response.url,
                                timeout=15_000,
                            ) as check_right_info:
                                export_button.click(timeout=15_000)
                            check_right_response = check_right_info.value
                            try:
                                check_right_body = check_right_response.text()
                            except Exception as exc:
                                check_right_body = f"<unable to read response body: {exc}>"
                            diagnostic_events["check_right_response"].append(
                                f"{check_right_response.status} {check_right_response.url}\n"
                                f"{check_right_body[:10000]}"
                            )
                            try:
                                parsed_right_check = json.loads(check_right_body)
                                if isinstance(parsed_right_check, dict):
                                    right_check = parsed_right_check
                            except json.JSONDecodeError:
                                pass
                        except Exception:
                            export_button.click(timeout=15_000)
                        update(
                            phase="waiting_download",
                            message="已点击“导出数据”，正在等待蝉妈妈生成并下载Excel；如页面出现确认弹窗，请手动确认。",
                            busy=True,
                            diagnostics_path=str(diagnostics_path),
                            **page_details(page),
                        )
                        page.wait_for_timeout(4_000)
                        if not downloaded:
                            if int(right_check.get("errCode") or 0) == 52000:
                                local_data = capture_visible_leaderboard(page)
                                local_data.update(
                                    {
                                        "title": f"蝉妈妈带货达人榜｜{category}｜{period_label}",
                                        "subtitle": (
                                            "蝉妈妈官方导出权限不足；本文件由主播发现 Agent "
                                            "根据账号当前可见榜单生成"
                                        ),
                                    }
                                )
                                destination = download_dir / safe_file_name(
                                    f"蝉妈妈_{category}_{period_label}_可见榜单.xlsx"
                                )
                                preview_path = build_local_leaderboard_xlsx(
                                    local_data, destination, state_path
                                )
                                downloaded.append(destination)
                                update(
                                    phase="downloaded",
                                    message=(
                                        "蝉妈妈官方导出权限不足，Agent已改为读取当前可见榜单，"
                                        f"并在本地生成Excel：{destination.name}"
                                    ),
                                    download_path=str(destination),
                                    preview_path=str(preview_path),
                                    official_export_error=str(right_check.get("errMsg") or ""),
                                    busy=False,
                                    imported=False,
                                )
                            else:
                                screenshot_path = state_path.with_name("after_export_click.png")
                                page.screenshot(path=str(screenshot_path), full_page=True)
                                diagnostics_path = save_export_diagnostics(
                                    page, state_path, diagnostic_events, export_button
                                )
                                update(
                                    phase="needs_action",
                                    message="已点击“导出数据”，但蝉妈妈没有开始下载。正在检查账号权益、页面弹窗或脚本错误。",
                                    busy=True,
                                    diagnostics_path=str(diagnostics_path),
                                    screenshot_path=str(screenshot_path),
                                    **page_details(page),
                                )
                    except Exception as exc:
                        debug_map = capture_page_map(page, state_path, "page_map_export_blocked.json")
                        screenshot_path = state_path.with_name("export_blocked.png")
                        page.screenshot(path=str(screenshot_path), full_page=True)
                        update(
                            phase="needs_action",
                            message=f"已进入目标榜单，但“导出数据”暂时无法点击：{exc}",
                            busy=True,
                            debug_map=str(debug_map),
                            screenshot_path=str(screenshot_path),
                            **page_details(page),
                        )
            else:
                route_maps: list[str] = []
                routes = [
                    ("creator", "https://www.chanmama.com/bloggerRank/"),
                    ("live_creator", "https://www.chanmama.com/bloggerRank/liveBlogger/"),
                    ("rising_creator", "https://www.chanmama.com/bloggerRank/bloggerRise/"),
                ]
                for route_name, route_url in routes:
                    page.goto(route_url, wait_until="domcontentloaded", timeout=60_000)
                    page.wait_for_timeout(3000)
                    route_maps.append(
                        str(capture_page_map(page, state_path, f"page_map_{route_name}.json"))
                    )
                page.goto(routes[1][1], wait_until="domcontentloaded", timeout=60_000)
                page.wait_for_timeout(2500)
                page.get_by_text("服饰内衣", exact=True).first.click(timeout=15_000)
                page.wait_for_timeout(1500)
                page.get_by_text("周榜", exact=True).first.click(timeout=15_000)
                page.wait_for_timeout(1500)
                route_maps.append(
                    str(capture_page_map(page, state_path, "page_map_selector_check.json"))
                )
                map_path = Path(route_maps[0])
                update(
                    phase="calibration_ready",
                    message="蝉妈妈页面结构已读取，正在校准达人榜路径",
                    mode="calibrate",
                    busy=True,
                    map_path=str(map_path),
                    route_maps=route_maps,
                    **page_details(page),
                )

            deadline = time.time() + 30 * 60
            while time.time() < deadline and not stop_path.exists() and not downloaded:
                time.sleep(1)
                if int(time.time()) % 10 == 0:
                    update(**page_details(page))

            stop_reason = stop_path.read_text(encoding="utf-8").strip() if stop_path.exists() else ""
            if args.mode == "login" and stop_reason == "login-complete":
                update(
                    phase="ready",
                    message="蝉妈妈专用登录状态已保存",
                    logged_in=True,
                    busy=False,
                    mode="",
                    **page_details(page),
                )
            elif not downloaded and stop_path.exists():
                update(
                    phase="ready" if initial_logged_in else "not_configured",
                    message="已取消本次蝉妈妈操作",
                    logged_in=initial_logged_in,
                    busy=False,
                    mode="",
                )
            elif not downloaded:
                update(phase="error", message="等待蝉妈妈操作超时，请重新尝试", busy=False, mode="")
            # A CDP-attached context belongs to the visible dedicated login
            # window. Disconnect without closing it; a context launched by
            # this worker is ours and should be closed normally.
            if attached_browser is None:
                context.close()
    except Exception as exc:
        update(phase="error", message=f"蝉妈妈专用浏览器运行失败：{exc}", busy=False)
        return 1
    finally:
        stop_path.unlink(missing_ok=True)
    return 0


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("mode", choices=["login", "export", "calibrate"])
    parser.add_argument("--profile-dir", required=True)
    parser.add_argument("--download-dir", required=True)
    parser.add_argument("--state-path", required=True)
    parser.add_argument("--stop-path", required=True)
    parser.add_argument("--start-url", required=True)
    parser.add_argument("--theme-id", type=int)
    parser.add_argument("--category", default="全部")
    parser.add_argument("--period", choices=["day", "week", "month"], default="day")
    raise SystemExit(run(parser.parse_args()))


if __name__ == "__main__":
    main()
