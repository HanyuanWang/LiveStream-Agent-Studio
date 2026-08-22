from __future__ import annotations

import csv
import io
import math
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


FIELD_ALIASES: dict[str, list[str]] = {
    "anchor_name": ["主播", "主播昵称", "达人", "达人昵称", "播主", "账号名称", "昵称", "达人名称"],
    "douyin_id": ["抖音号", "达人抖音号", "账号id", "达人id", "unique_id"],
    "profile_url": ["主页链接", "抖音主页", "主播主页", "达人主页", "profile_url", "主页地址", "直播间链接"],
    "analysis_url": ["蝉妈妈详情页", "达人详情页", "分析链接", "analysis_url"],
    "followers": ["粉丝数", "粉丝", "达人粉丝", "当前粉丝数"],
    "category": ["品类", "类目", "主营类目", "达人分类", "行业", "带货类目"],
    "estimated_gmv_text": ["预估销售额", "直播销售额", "直播销售额(元)", "销售额", "gmv", "预估gmv", "场均销售额"],
    "gmv_index": ["销售额指数", "直播销售额指数", "gmv指数"],
    "sales_volume_text": ["预估销量", "直播销量", "直播销量(件)", "销量", "销售量", "订单量"],
    "sales_index": ["销量指数", "直播销量指数"],
    "gpm": ["gpm", "千次观看成交金额", "千次观看成交"],
    "uv_value": ["uv价值", "uv价值指数", "uv value", "点击价值"],
    "avg_online": ["平均在线", "平均在线人数", "场均在线", "平均观看人数", "人气峰值"],
    "duration_hours": ["直播时长", "时长", "场均时长", "直播时长(小时)"],
    "sessions_7d": ["近7日场次", "7日直播场次", "直播场次", "近7天直播场次"],
    "stability": ["稳定性", "稳定度", "直播稳定性"],
    "account_type": ["账号类型", "主播类型", "达人类型", "直播类型"],
    "title": ["直播标题", "标题", "直播间标题"],
    "products": ["商品", "主营商品", "带货商品", "热销商品", "商品名称"],
}


def normalize_header(value: Any) -> str:
    return re.sub(r"[\s_（）()\-—/]+", "", str(value or "").strip().lower())


ALIAS_LOOKUP = {
    normalize_header(alias): field
    for field, aliases in FIELD_ALIASES.items()
    for alias in aliases
}


def parse_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value) if math.isfinite(float(value)) else None
    text = str(value).strip().replace(",", "").replace("￥", "").replace("¥", "")
    if not text or text in {"-", "--", "暂无", "未知"}:
        return None
    range_match = re.search(r"([\d.]+)\s*([万亿wk]?)\s*[-~～至]\s*([\d.]+)\s*([万亿wk]?)", text, re.I)
    if range_match:
        left = _scale(float(range_match.group(1)), range_match.group(2) or range_match.group(4))
        right = _scale(float(range_match.group(3)), range_match.group(4) or range_match.group(2))
        return (left + right) / 2
    match = re.search(r"-?[\d.]+", text)
    if not match:
        return None
    unit_match = re.search(r"[万亿wk]", text, re.I)
    return _scale(float(match.group()), unit_match.group() if unit_match else "")


def _scale(number: float, unit: str) -> float:
    unit = unit.lower()
    if unit in {"万", "w"}:
        return number * 10_000
    if unit == "亿":
        return number * 100_000_000
    if unit == "k":
        return number * 1_000
    return number


def parse_duration_hours(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if re.fullmatch(r"\d{1,2}:\d{2}(?::\d{2})?", text):
        parts = [int(part) for part in text.split(":")]
        if len(parts) == 2:
            return parts[0] + parts[1] / 60
        return parts[0] + parts[1] / 60 + parts[2] / 3600
    hours = re.search(r"([\d.]+)\s*(?:小时|h)", text, re.I)
    minutes = re.search(r"([\d.]+)\s*(?:分钟|min)", text, re.I)
    if hours or minutes:
        return (float(hours.group(1)) if hours else 0) + (float(minutes.group(1)) / 60 if minutes else 0)
    return parse_number(text)


def _is_range(value: Any) -> bool:
    return isinstance(value, str) and bool(re.search(r"[-~～至+]|\b以上\b", value))


def _metric_fields(value: Any) -> tuple[float | None, str]:
    if value in (None, ""):
        return None, ""
    text = str(value).strip()
    return (None if _is_range(value) else parse_number(value)), text


def _rows_from_xlsx(content: bytes) -> list[list[Any]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def _hyperlinks_from_xlsx(content: bytes) -> dict[tuple[int, int], str]:
    workbook = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    sheet = workbook.active
    links: dict[tuple[int, int], str] = {}
    for row in sheet.iter_rows():
        for cell in row:
            target = getattr(getattr(cell, "hyperlink", None), "target", None)
            if target:
                links[(cell.row - 1, cell.column - 1)] = str(target)
    return links


def _rows_from_csv(content: bytes) -> list[list[str]]:
    text = None
    for encoding in ["utf-8-sig", "gb18030", "utf-8"]:
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("无法识别CSV编码")
    return list(csv.reader(io.StringIO(text)))


def read_tabular(file_name: str, content: bytes) -> list[list[Any]]:
    suffix = Path(file_name).suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _rows_from_xlsx(content)
    if suffix in {".csv", ".txt"}:
        return _rows_from_csv(content)
    raise ValueError("只支持 .xlsx、.xlsm、.csv 或 .txt 榜单")


def _find_header(rows: list[list[Any]]) -> tuple[int, dict[int, str]]:
    best_index = -1
    best_mapping: dict[int, str] = {}
    for row_index, row in enumerate(rows[:20]):
        mapping: dict[int, str] = {}
        for column_index, value in enumerate(row):
            field = ALIAS_LOOKUP.get(normalize_header(value))
            if field:
                mapping[column_index] = field
        if len(mapping) > len(best_mapping):
            best_index, best_mapping = row_index, mapping
    if best_index < 0 or "anchor_name" not in best_mapping.values():
        raise ValueError("没有找到主播名称列；请确保表格包含“主播/达人昵称/播主”等表头")
    return best_index, best_mapping


def parse_leaderboard(file_name: str, content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    rows = read_tabular(file_name, content)
    hyperlinks = _hyperlinks_from_xlsx(content) if Path(file_name).suffix.lower() in {".xlsx", ".xlsm"} else {}
    if not rows:
        return [], ["文件为空"]
    header_index, mapping = _find_header(rows)
    candidates: list[dict[str, Any]] = []
    warnings: list[str] = []
    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        raw = {str(rows[header_index][index] or f"列{index + 1}"): value for index, value in enumerate(row) if value not in (None, "")}
        data = {field: row[index] if index < len(row) else None for index, field in mapping.items()}
        anchor_name = str(data.get("anchor_name") or "").strip()
        if not anchor_name:
            continue
        douyin_id = str(data.get("douyin_id") or "").strip()
        profile_url = str(data.get("profile_url") or "").strip()
        anchor_column = next((index for index, field in mapping.items() if field == "anchor_name"), -1)
        anchor_link = hyperlinks.get((row_number - 1, anchor_column), "") if anchor_column >= 0 else ""
        analysis_url = str(data.get("analysis_url") or "").strip()
        if not analysis_url and (
            "chanmama.com" in anchor_link or anchor_link.startswith("/bloggerRank/")
        ):
            analysis_url = anchor_link
        if not profile_url and anchor_link and not analysis_url:
            profile_url = anchor_link
        source_key = douyin_id or profile_url or anchor_name
        estimated_gmv, estimated_gmv_text = _metric_fields(data.get("estimated_gmv_text"))
        sales_volume, sales_volume_text = _metric_fields(data.get("sales_volume_text"))
        candidates.append(
            {
                "source_key": source_key,
                "anchor_name": anchor_name,
                "douyin_id": douyin_id,
                "profile_url": profile_url,
                "analysis_url": analysis_url,
                "followers": parse_number(data.get("followers")),
                "category": str(data.get("category") or "").strip(),
                "estimated_gmv": estimated_gmv,
                "estimated_gmv_text": estimated_gmv_text,
                "gmv_index": parse_number(data.get("gmv_index")),
                "sales_volume": sales_volume,
                "sales_volume_text": sales_volume_text,
                "sales_index": parse_number(data.get("sales_index")),
                "gpm": parse_number(data.get("gpm")),
                "uv_value": parse_number(data.get("uv_value")),
                "avg_online": parse_number(data.get("avg_online")),
                "duration_hours": parse_duration_hours(data.get("duration_hours")),
                "sessions_7d": parse_number(data.get("sessions_7d")),
                "stability": parse_number(data.get("stability")),
                "account_type": str(data.get("account_type") or "").strip(),
                "title": str(data.get("title") or "").strip(),
                "products": str(data.get("products") or "").strip(),
                "raw_data": raw,
                "row_number": row_number,
            }
        )
    if not candidates:
        warnings.append("识别到表头，但没有读到有效主播行")
    recognized = set(mapping.values())
    missing = [field for field in ["followers", "estimated_gmv_text", "gpm", "uv_value", "profile_url"] if field not in recognized]
    if missing:
        warnings.append("未找到部分评分字段：" + "、".join(missing))
    return candidates, warnings
